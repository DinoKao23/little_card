import streamlit as st
import polars as pl
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.styles import Border, Side
import io




def automize_openpyxl(wb, pu_code, rubber_code, fabric_code, liner_code, uv_code, liner_texture_code, sheet_name = None):
    pu_thickness = pu_df.filter(pl.col("SurrogateKey") == pu_code).select("THICKNESS_mm").item()
    rubber_thickness = rubber_df.filter(pl.col("SurrogateKey") == rubber_code).select("THICKNESS_mm").item()
    fabric_thickness = fabric_df.filter(pl.col("SurrogateKey") == fabric_code).select("THICKNESS_mm").item()
    liner_thickness = liner_df.filter(pl.col("SurrogateKey") == liner_code).select("THICKNESS_mm").item()
    uv_thickness = uv_print_df.filter(pl.col("SurrogateKey") == uv_code).select("THICKNESS_mm").item()

    total_thickness = pu_thickness + rubber_thickness + fabric_thickness + liner_thickness + uv_thickness

    pu_weight = pu_df.filter(pl.col("SurrogateKey") == pu_code).select("WEIGHT_gsm").item()
    rubber_weight = rubber_df.filter(pl.col("SurrogateKey") == rubber_code).select("WEIGHT_gsm").item()
    fabric_weight = fabric_df.filter(pl.col("SurrogateKey") == fabric_code).select("WEIGHT_gsm").item()
    liner_weight = liner_df.filter(pl.col("SurrogateKey") == liner_code).select("WEIGHT_gsm").item()
    uv_weight = uv_print_df.filter(pl.col("SurrogateKey") == uv_code).select("WEIGHT_gsm").item()

    total_weight = pu_weight + rubber_weight + fabric_weight + liner_weight + uv_weight

    pu_weight_percent = round(pu_weight / total_weight * 100)
    rubber_weight_percent = round(rubber_weight / total_weight * 100)
    fabric_weight_percent = round(fabric_weight / total_weight * 100)
    liner_weight_percent = round(liner_weight / total_weight * 100)
    uv_weight_percent = round(uv_weight / total_weight * 100)
    total_weight_percentage = pu_weight_percent + rubber_weight_percent + fabric_weight_percent + liner_weight_percent + uv_weight_percent

    if total_weight_percentage < 100:
        rubber_weight_percent = rubber_weight_percent + (100 - total_weight_percentage)
    elif total_weight_percentage > 100:
        rubber_weight_percent = rubber_weight_percent - (total_weight_percentage - 100)

    uv_pu_weight_percent = uv_weight_percent + pu_weight_percent

    biobased = 0

    if fabric_df.filter(pl.col("SurrogateKey") == fabric_code).select("unnatural").item() == 1 & liner_df.filter(pl.col("SurrogateKey") == liner_code).select("unnatural").item() == 1:
        biobased = round(pu_weight_percent*0.94, 2)
    elif fabric_df.filter(pl.col("SurrogateKey") == fabric_code).select("unnatural").item() == 0 & liner_df.filter(pl.col("SurrogateKey") == liner_code).select("unnatural").item() == 1:
        biobased = round(pu_weight_percent*0.94 + fabric_weight_percent, 2)
    elif fabric_df.filter(pl.col("SurrogateKey") == fabric_code).select("unnatural").item() == 1 & liner_df.filter(pl.col("SurrogateKey") == liner_code).select("unnatural").item() == 0:
        biobased = round(pu_weight_percent*0.94 + liner_weight_percent, 2)
    else :
        biobased = round(pu_weight_percent*0.94 + fabric_weight_percent + liner_weight_percent, 2)
    
    # Check if the sheet exists and delete it
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    # Create a new worksheet
    sht = wb.create_sheet(title=sheet_name)

    # Set background color (OpenPyXL only supports cell-by-cell coloring)
    fill = PatternFill(start_color="D7E4CA", end_color="D7E4CA", fill_type="solid")  # Light green

    for row in sht.iter_rows(min_row=1, max_row=50, min_col=1, max_col=26):  # A1:Z50
        for cell in row:
            cell.fill = fill

    # Define alignment with text wrapping
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Define data
    liner_texture = liner_texture_df.filter(pl.col('SurrogateKey') == 1).select(pl.col('Texture No.')).item()
    liner_number = liner_df.filter(pl.col('SurrogateKey') == 1).select(pl.col('Natuura_key')).item()
    rubber = rubber_df.filter(pl.col('SurrogateKey') == 1).select(pl.col('Short Name')).item()
    fabric = fabric_df.filter(pl.col('SurrogateKey') == 1).select(pl.col('Short Name')).item()
    liner = liner_df.filter(pl.col('SurrogateKey') == 1).select(pl.col('Short Name')).item()
    uv = uv_print_df.filter(pl.col('SurrogateKey') == 1).select(pl.col('Short Name')).item()

    # Merge cells and apply data
    sht["A1"] = "Company\nTORY BURCH"
    sht["A1"].alignment = wrap_alignment
    sht["B1"] = "BRAND\nLUCID MAT™"
    sht["B1"].alignment = wrap_alignment
    sht["C1"] = "STYLE\nPatent"
    sht["C1"].alignment = wrap_alignment
    sht["D1"] = "DATE\n2025-01-20"
    sht["D1"].alignment = wrap_alignment

    sht["A2"] = f"FINISH UV printing/glossy"
    sht["A2"].alignment = wrap_alignment
    sht["B2"] = f"TEXTURE\n{liner_texture} "
    sht["B2"].alignment = wrap_alignment
    sht["C2"] = f"LINER NO.\n{liner_number}"
    sht["C2"].alignment = wrap_alignment
    sht["D2"] = f"WEIGHT\n{total_weight}gsm ± 20%"
    sht["D2"].alignment = wrap_alignment


    sht.merge_cells("A3:C3")
    sht["A3"] = "COLOR\n"
    sht["A3"].alignment = wrap_alignment

    sht["D3"] = f"THICKNESS\n{total_thickness}mm ± 0.2mm"
    sht["D3"].alignment = wrap_alignment

    sht.merge_cells("A4:C9")
    sht.merge_cells("D4:D9")

    if uv != 'NA':
        sht["A4"] = f"""COMPOSITION
{rubber}: 
{fabric}: 
{liner}:
DRY POLYURETHANE 
+TOP COATING & UV PRINT:"""
        sht["D4"] = f"""Weight %
{rubber_weight_percent}%
{fabric_weight_percent}%
{liner_weight_percent}%

{uv_pu_weight_percent}%"""
    else:
        sht["A4"] = f"""COMPOSITION
{rubber}:
{fabric}:
{liner}:
DRY POLYURETHANE:"""
        sht["D4"] = f"""
{rubber_weight_percent}%
{fabric_weight_percent}%
{liner_weight_percent}%

{pu_weight_percent}%"""

    sht["A4"].alignment = wrap_alignment
    wrap_left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    sht["D4"].alignment = wrap_left_alignment


    sht["A11"] = "ESTIMATED BIOBASED:"
    sht["D11"] = f"{biobased}%"
    sht["D11"].alignment = wrap_alignment

    sht.merge_cells("A12:D15")
    sht["A12"] = """MOQ / MCQ, SIZE, LEAD TIME
    MOQ – 300m ; MCQ – 300m
    Bulk Lead Time – 2 Months (with In-Stock Liner)
    Usable Width – 1.32m (Up to 30m Per Roll)
    """
    sht["A12"].alignment = wrap_alignment

    sht.merge_cells("A16:D23")
    sht["A16"] = """OPTIONS
    Overall Thickness – 1.0mm – 2.0mm
    Pigmentation – Pantone, Custom
    Variety of Continuous Textures
    Liners – Biobased - 100% GOTS Organic Cotton, 100% Cotton, 100% TENCEL™ (Canvas or Interlock)
            Recycled - 100% GRS Post-Consumer Recycled Polyester
            (Knit or Suede)
    """
    sht["A16"].alignment = wrap_alignment


    # Apply Borders
    border_style = Side(style="medium")  # xlMedium equivalent in OpenPyXL
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    for row in range(1, 24):  # A1:D23
        for col in ["A", "B", "C", "D"]:
            sht[f"{col}{row}"].border = border

    side = Side(border_style=None)
    no_border = openpyxl.styles.borders.Border(
        left=side, 
        right=side, 
        top=side, 
        bottom=side,
    )

    fix_border = openpyxl.styles.borders.Border(
        left=side, 
        top=side, 
        bottom=side,
    )

    fix_border_change = openpyxl.styles.borders.Border(
        left=side, 
        top=side, 
        bottom=side,
        right = border_style
    )

    for row in range(10, 12):  # A1:D23
        for col in ["A", "B", "C",]:
            sht[f"{col}{row}"].border = no_border

    for row in range(4, 12):  # A1:D23
        for col in ["A","B", "C","D"]:
            sht[f"{col}{row}"].border = fix_border

    for row in range(4, 12):  # A1:D23
        for col in ["D"]:
            sht[f"{col}{row}"].border = fix_border_change

    # Set column widths
    sht.column_dimensions["A"].width = 12.56
    sht.column_dimensions["B"].width = 11.44
    sht.column_dimensions["C"].width = 11.44
    sht.column_dimensions["D"].width = 15.22



# Example file path (replace with your file path)
file_path = "data.xlsx"

# Load DataFrames
pu_df = pl.read_excel(file_path, sheet_name="PU")
rubber_df = pl.read_excel(file_path, sheet_name="Rubber")
fabric_df = pl.read_excel(file_path, sheet_name="Fabric")
liner_df = pl.read_excel(file_path, sheet_name="Liner")
liner_texture_df = pl.read_excel(file_path, sheet_name="liner_texture")
uv_print_df = pl.read_excel(file_path, sheet_name="UV_print")

fabric_df = fabric_df.with_columns(
    unnatural = (pl.col("Short Name").str.contains("PET", strict=False)).cast(pl.Int8())
)
liner_df = liner_df.with_columns(
    unnatural = (pl.col("Short Name").str.contains("PET", strict=False)).cast(pl.Int8())
)

# Define dropdown data
dropdown_dataframes = [pu_df, rubber_df, fabric_df, liner_df, uv_print_df, liner_texture_df]
dropdown_columns = ["Short Name", "Short Name", "Short Name", "Short Name", "Short Name", "Surface Texture & Finish"]
dropdown_titles = [
    "選擇 PU", 
    "選擇 Rubber", 
    "選擇 Fabric", 
    "選擇 Liner", 
    "選擇 UV Print", 
    "選擇 Liner Texture"
]

# ✅ **Initialize Session State**
if "user_selections" not in st.session_state:
    st.session_state.user_selections = [None] * 6
if "sheet_name" not in st.session_state:
    st.session_state.sheet_name = ""

# Function to map selection to SurrogateKey
def map_selection_to_code(selected_value, df, col_name):
    try:
        return df.filter(pl.col(col_name) == selected_value).select("SurrogateKey").item()
    except:
        return None  # Return None if not found

# Streamlit UI
st.title("綠色小卡製作")

num_options = len(dropdown_dataframes)  # Automatically set size based on dropdowns
if "user_selections" not in st.session_state:
    st.session_state.user_selections = [None] * num_options

# 🎯 **Display all selections on the same page**
for i, df in enumerate(dropdown_dataframes):
    col_name = dropdown_columns[i]
    title = dropdown_titles[i]

    selected_value = st.selectbox(title, df[col_name].to_list(), key=f"selection_{i}")

    # Save mapped code to session state
    st.session_state.user_selections[i] = map_selection_to_code(selected_value, df, col_name)

# 📌 **Sheet Name Input**
sheet_name = st.text_input("輸入表單名稱", value=st.session_state.get("sheet_name", ""))
st.session_state.sheet_name = sheet_name
if "workbook" not in st.session_state:
    st.session_state.workbook = openpyxl.Workbook()
    # Remove default sheet created by openpyxl
    default_sheet = st.session_state.workbook.active
    st.session_state.workbook.remove(default_sheet)
    
if "sheet_count" not in st.session_state:
    st.session_state.sheet_count = 0  # Track number of sheets added

if st.button("建立表單"):
    if None in st.session_state.user_selections or not st.session_state.sheet_name:
        st.error("⚠️ Please make all selections before running automation.")
    else:
        # Increment sheet count to avoid duplicate names
        st.session_state.sheet_count += 1
        sheet_name = f"{st.session_state.sheet_name}_{st.session_state.sheet_count}"
        
        # Run automation function and append to existing workbook
        automize_openpyxl(
            st.session_state.workbook,
            pu_code=st.session_state.user_selections[0],
            rubber_code=st.session_state.user_selections[1],
            fabric_code=st.session_state.user_selections[2],
            liner_code=st.session_state.user_selections[3],
            uv_code=st.session_state.user_selections[4],
            liner_texture_code=st.session_state.user_selections[5],
            sheet_name=sheet_name
        )

        st.success(f"🎉 Automation Completed! Sheet '{sheet_name}' added to the Excel file.")

if st.button("完成表單選擇"):
    if st.session_state.sheet_count == 0:
        st.error("⚠️ No sheets to download. Run automation first.")
    else:
        # Save workbook to BytesIO
        output = io.BytesIO()
        st.session_state.workbook.save(output)
        output.seek(0)  # Reset stream position

        st.download_button(
            label="📥 下載文件",
            data=output,
            file_name="小卡.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# 🔄 **Restart Button**
if st.button("清除已存取表單"):
    st.session_state.user_selections = [None] * num_options
    st.session_state.sheet_name = ""

    # Reinitialize the workbook (starts fresh)
    st.session_state.workbook = openpyxl.Workbook()
    default_sheet = st.session_state.workbook.active
    st.session_state.workbook.remove(default_sheet)

    # Reset sheet count
    st.session_state.sheet_count = 0  

    st.rerun()  # Rerun app to reflect changes
